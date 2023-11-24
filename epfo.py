from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from selenium.common.exceptions import NoSuchElementException
import time
import re
import math
import os

# workbook setup
workbook = Workbook()
worksheet = workbook.active
begin = 1

# Setup driver and fetch website
DRIVER_PATH = r"chromedriver"

# print epfo ascii art
print("""
███████╗██████╗░███████╗░█████╗░
██╔════╝██╔══██╗██╔════╝██╔══██╗
█████╗░░██████╔╝█████╗░░██║░░██║
██╔══╝░░██╔═══╝░██╔══╝░░██║░░██║
███████╗██║░░░░░██║░░░░░╚█████╔╝
╚══════╝╚═╝░░░░░╚═╝░░░░░░╚════╝░
""")
print("**** INITIALIZING DRIVER ****")
driver = webdriver.Chrome(executable_path=DRIVER_PATH)
url="https://unifiedportal-epfo.epfindia.gov.in/publicPortal/no-auth/misReport/home/loadEstSearchHome"
driver.get(url)
time.sleep(3)

COMPANY_NAME = input("Enter company name: ")

name_of_establishment = driver.find_element(By.ID,"estName")
name_of_establishment.clear()
name_of_establishment.send_keys(COMPANY_NAME)

captacha = driver.find_element(By.ID,"captcha")
captacha.clear()
captacha_value = input("Enter Captacha: ")
captacha.send_keys(captacha_value)

search_button = driver.find_element(By.ID,"searchEmployer")
search_button.click()
time.sleep(4)
# count the number of rows in the table
def row_count():
    # count the number of rows in the table
    table = driver.find_element(By.ID,"example")
    rows = table.find_elements(By.TAG_NAME, "tr")
    print("Number of rows in the table: ",len(rows))
    return len(rows)

# function to click the link in the table
def table_click(row):
    links = driver.find_elements(By.XPATH, '//*[@id="example"]/tbody/tr['+str(row)+']/td[5]/a')
    if links:
        links[0].click()

# add data to the table
def create_tables(start_val):
    html_content = driver.page_source

    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#tablecontainer3 table')

    rows = table.find_all('tr')

    for row_index, row in enumerate(rows, start=start_val):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)

    table = soup.select_one('#tablecontainer4 table')

    rows = table.find_all('tr')

    for row_index, row in enumerate(rows, start=start_val + 6):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
    try:
        pan_status = driver.find_element(By.XPATH,'//*[@id="tablecontainer5"]/div[2]/table/tbody/tr[2]/td[2]')

        worksheet.cell(row=start_val + 13,column=1,value="PAN STATUS:")
        worksheet.cell(row=start_val + 13,column=2,value=pan_status.text)
    except NoSuchElementException:
        print("No PAN Status")
    try:
        esic_code = driver.find_element(By.XPATH, '//*[@id="tablecontainer12"]/table/tbody/tr[3]/td[2]')
        worksheet.cell(row=start_val + 14,column=1,value="ESIC Code:")
        worksheet.cell(row=start_val + 14,column=2,value=esic_code.text)
    except NoSuchElementException:
        print("No ESIC Code")

    row_color = '00FF00'
    fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

    row_number = start_val + 15
    row = worksheet[row_number]

    for cell in row:
        cell.fill = fill

def payments_table(begin)->int:
    html_content = driver.page_source

    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#table_pop_up')

    rows = table.find_all('tr')

    for row_index, row in enumerate(rows, start=begin):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
        begin += 1
    return begin
try:
    # grab each cell
    total_records = driver.find_element(By.XPATH,'//*[@id="collapseTwo"]/div[1]/div')
    pattern = r"\d+"
    matches = re.findall(pattern , total_records.text)
    if matches:
        number = int(matches[0])
        print(number)

    number = math.ceil(number / 10)
    print(number)
    for i in range(1,number+1):
        row_len = row_count()
        for row in range(1, row_len):
            table_click(row)
            time.sleep(5)
            create_tables(begin)
            begin = begin + 17
            payments_link = driver.find_element(By.XPATH,'//*[@id="tablecontainer3"]/div/a/u')
            payments_link.click()
            time.sleep(5)

            driver.switch_to.window(driver.window_handles[1])
            row_number = begin + 1
            try:
                pages_in_payments_text = driver.find_element(By.XPATH,'//*[@id="table_pop_up_info"]')
                pattern = r"(\d+)$"

                nos_pages_in_payments = re.search(pattern,pages_in_payments_text.text)
                if nos_pages_in_payments:
                    number_of_pays = nos_pages_in_payments.group(1)
                    print("Number of payments pages: "+number_of_pays)
                for nos in range(1,int(number_of_pays)+1):
                    begin = payments_table(begin)
                    row_number = begin + 1
                    next_button = driver.find_element(By.XPATH,'//*[@id="table_pop_up_next"]')
                    next_button.click()
            except NoSuchElementException:
                print("No payments")
            row_color = '00FF00'
            fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            row = worksheet[row_number]

            for cell in row:
                cell.fill = fill
            time.sleep(5)
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)
            begin = begin + 2

        next = driver.find_element(By.XPATH,'//*[@id="example_next"]')
        next.click()
        time.sleep(3)

    # click on the same pan link
    # try:
    #     same_pan = driver.find_element(By.XPATH,'//*[@id="tablecontainer11"]/div/a')
    #     same_pan.click()
    # except NoSuchElementException:
    #     print("No same pan")

    # save file in a folder
    gwd = os.getcwd()
    path = os.path.join(gwd,"Files")
    if not os.path.exists(path):
        os.mkdir(path)
    os.chdir(path)
    workbook.save(filename=str(COMPANY_NAME + ".xlsx"))
except NoSuchElementException:
    print("No data found")