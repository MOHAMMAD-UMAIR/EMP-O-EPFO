from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import time
import datetime
import re
import math

# workbook setup
workbook = Workbook()
worksheet = workbook.active
begin = 1

# Setup driver and fetch website
DRIVER_PATH = r"chromedriver"

# print epfo ascii art
print("""
███████╗██████╗░███████╗░█████╗░
██╔════╝██╔══██╗██╔════╝██╔══██╗
█████╗░░██████╔╝█████╗░░██║░░██║
██╔══╝░░██╔═══╝░██╔══╝░░██║░░██║
███████╗██║░░░░░██║░░░░░╚█████╔╝
╚══════╝╚═╝░░░░░╚═╝░░░░░░╚════╝░
""")
print("**** INITIALIZING DRIVER ****")
driver = webdriver.Chrome(executable_path=DRIVER_PATH)
url="https://unifiedportal-epfo.epfindia.gov.in/publicPortal/no-auth/misReport/home/loadEstSearchHome"
driver.get(url)
time.sleep(3)

i = input("Search by Establishment Code or Company Name? (E/C):")
if i == "C":
    COMPANY_NAME = input("Enter company name: ")
    name_of_establishment = driver.find_element(By.ID,"estName")
    name_of_establishment.clear()
    name_of_establishment.send_keys(COMPANY_NAME)
elif i == "E":
    EST_ID = input("Establishment Code: ")
    COMPANY_NAME = input("Enter company name: ")
    est = driver.find_element_by_id("estCode")
    est.clear()
    est.send_keys(EST_ID)
else:
    print("Invalid input. Exiting...")
    exit()

captacha = driver.find_element(By.ID,"captcha")
captacha.clear()
captacha_value = input("Enter Captacha: ")
captacha.send_keys(captacha_value)

search_button = driver.find_element(By.ID,"searchEmployer")
search_button.click()
time.sleep(5)

date1 = datetime.datetime(2022,6,1)

workbook = Workbook()
worksheet = workbook.active
begin = 1

def row_count():
    # count the number of rows in the table
    table = driver.find_element(By.ID,"example")
    rows = table.find_elements(By.TAG_NAME, "tr")
    print("Number of rows in the table: ",len(rows))
    return len(rows)

def table_click(row):
    links = driver.find_elements(By.XPATH, '//*[@id="example"]/tbody/tr['+str(row)+']/td[5]/a')
    if links:
        links[0].click()

def table_click_2(row):
    E_ID = driver.find_elements(By.XPATH, '//*[@id="table_pop_up"]/tbody/tr['+str(row)+']/td[1]/a')
    if E_ID:
        E_ID[0].click()
        return True

def row_count_2():
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#table_pop_up_wrapper table')
    rows = table.find_all('tr')
    return len(rows)

def payments_table(begin)->int:
    html_content = driver.page_source

    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#table_pop_up')

    rows = table.find_all('tr')

    for row_index, row in enumerate(rows, start=begin):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            if col_index == 2:
                payment_date = datetime.datetime.strptime(cell_value,'%d-%b-%Y %H:%M:%S')
            if col_index == 4:
                wage_date = datetime.datetime.strptime(cell_value, '%b-%y')
                wage_month = wage_date.month
                wage_year = wage_date.year
                wage_day = wage_date.day
                if  wage_date >=  date1:
                    cell_value = wage_date.strftime("%b-%y")
                    worksheet.cell(row=row_index, column=col_index, value=cell_value)
                    next_month = (wage_month % 12) + 1
                    payment_due_date = datetime.datetime(wage_year, next_month, 15)
                    if payment_due_date < payment_date:
                        worksheet.cell(row=row_index, column=7,value="DELAYED")
                    else:
                        worksheet.cell(row=row_index, column=7,value="-")
                else:
                    begin -= 1
                    break
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
        begin += 1
    return begin-1

def create_tables(start_val,xpath_1,xpath_2,xpath_3,xpath_4):
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    bold_font = Font(bold=True)

    table = soup.select_one(xpath_1)
    rows = table.find_all('tr')
    worksheet.cell(row=start_val,column=1,value="Validity Status-").font = bold_font
    for row_index, row in enumerate(rows, start=start_val+1):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)


    table = soup.select_one(xpath_2)
    rows = table.find_all('tr')
    worksheet.cell(row=start_val + 6,column=1,value="Establishment Status-").font = bold_font
    for row_index, row in enumerate(rows, start=start_val + 7):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)

    table = soup.select_one(xpath_3)
    rows = table.find_all('tr')

    worksheet.cell(row=start_val + 12,column=1,value="Establishment Details-").font = bold_font
    for row_index, row in enumerate(rows, start=start_val + 13):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            if cell_value == 'VERIFIED':
                fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                worksheet.cell(row=row_index, column=col_index).fill = fill
            elif cell_value == 'REJECTED':
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                worksheet.cell(row=row_index, column=col_index).fill = fill
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
    
    table = soup.select_one(xpath_4)
    rows = table.find_all('tr')

    worksheet.cell(row=start_val + 23,column=1,value="Additional Information-").font = bold_font
    for row_index, row in enumerate(rows, start=start_val + 24):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)


    row_color = 'ADD8E6'
    fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

    row_number = start_val + 32
    row = worksheet[row_number]

    for cell in row:
        cell.fill = fill

def heading(begin):
    bold_font = Font(bold=True)
    worksheet.cell(row=begin,column=1,value="TRRN").font = bold_font
    worksheet.cell(row=begin,column=2,value="Date of Credit").font = bold_font
    worksheet.cell(row=begin,column=3,value="Amount").font = bold_font
    worksheet.cell(row=begin,column=4,value="Wage Month").font = bold_font
    worksheet.cell(row=begin,column=5,value="No. of Employees").font = bold_font
    worksheet.cell(row=begin,column=6,value="ECR").font = bold_font
    worksheet.cell(row=begin,column=7,value="DELAY").font = bold_font

def header_2(begin):
    bold_font = Font(bold=True)
    worksheet.cell(row=begin,column=1,value="E_ID").font = bold_font
    worksheet.cell(row=begin,column=2,value="E_Name").font = bold_font
    worksheet.cell(row=begin,column=3,value="Address").font = bold_font
    worksheet.cell(row=begin,column=4,value="Office Name").font = bold_font

def same_pan_company(start_val)->int:
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#table_pop_up_wrapper table')

    rows = table.find_all('tr')
    
    for row_index, row in enumerate(rows, start=start_val):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
        start_val += 1

    return start_val

def count_pages()->int:
    total_pages = driver.find_element(By.XPATH,'//*[@id="table_pop_up_info"]')
    page_count = int(total_pages.text.split()[-1])
    print("Number of pages:",page_count)
    return page_count

def same_pan_company(start_val)->int:
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.select_one('#table_pop_up_wrapper table')

    rows = table.find_all('tr')
    
    for row_index, row in enumerate(rows, start=start_val):
        columns = row.find_all('td')
        for col_index, column in enumerate(columns, start=1):
            cell_value = column.get_text()
            worksheet.cell(row=row_index, column=col_index, value=cell_value)
        start_val += 1

    return start_val

total_records = driver.find_element(By.XPATH,'//*[@id="collapseTwo"]/div[1]/div')
pattern = r"\d+"
matches = re.findall(pattern , total_records.text)
if matches:
    number = int(matches[0])
    print(number)

number = math.ceil(number / 10)
print(number)
# for i in range(1,number+1):
for i in range(1,2):
    row_len = 2
    for row in range(1, row_len):
        table_click(row)
        time.sleep(30)
        create_tables(begin,'#tablecontainer3 table','#tablecontainer4 table','#tablecontainer5 table','#tablecontainer12 table')
        begin = begin + 33
        payments_link = driver.find_element(By.XPATH,'//*[@id="tablecontainer3"]/div/a/u')
        payments_link.click()
        time.sleep(10)

        driver.switch_to.window(driver.window_handles[1])
        pages_in_payments_text = driver.find_element(By.XPATH,'//*[@id="table_pop_up_info"]')
        pattern = r"(\d+)$"

        nos_pages_in_payments = re.search(pattern,pages_in_payments_text.text)
        if nos_pages_in_payments:
            number_of_pays = nos_pages_in_payments.group(1)
            print("Number of payment pages:",number_of_pays)
            heading(begin)
        for nos in range(1,int(number_of_pays)+1):
            begin = payments_table(begin)
            row_number = begin + 1
            next_button = driver.find_element(By.XPATH,'//*[@id="table_pop_up_next"]')
            next_button.click()
        row_color = 'ADD8E6'
        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        row = worksheet[row_number]

        for cell in row:
            cell.fill = fill
        time.sleep(1)
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(4)
        begin = begin + 1

    # next = driver.find_element(By.XPATH,'//*[@id="example_next"]')
    # next.click()
    # time.sleep(3)

same_pan = driver.find_element(By.XPATH,'//*[@id="tablecontainer11"]/div/a')
same_pan.click()
time.sleep(5)

driver.switch_to.window(driver.window_handles[1])

begin = same_pan_company(begin)

pages = count_pages()
for i in range(1,pages+1):
    num_of_rows = row_count_2()
    for row in range(1,num_of_rows+1):
        table_click_2(row)
        time.sleep(15)
        create_tables(begin,'#tbPopUp_3 table','#tbPopUp_4 table','#tbPopUp_5 table','#tbPopUp_12 table')
        begin += 33

        payments_link = driver.find_element(By.XPATH,'//*[@id="tbPopUp_3"]/div/a/u')
        payments_link.click()
        time.sleep(10)

        driver.switch_to.window(driver.window_handles[2])

        pages_in_payments_text = driver.find_element(By.XPATH,'//*[@id="table_pop_up_info"]')
        pattern = r"(\d+)$"
        nos_pages_in_payments = re.search(pattern,pages_in_payments_text.text)
        if nos_pages_in_payments:
            number_of_pays = nos_pages_in_payments.group(1)
            print("Number of payment pages:",number_of_pays)
            heading(begin)
        for nos in range(1,int(number_of_pays)+1):
            begin = payments_table(begin)
            row_number = begin + 1
            next_button = driver.find_element(By.XPATH,'//*[@id="table_pop_up_next"]')
            next_button.click()
        row_color = 'ADD8E6'
        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        row = worksheet[row_number]

        for cell in row:
            cell.fill = fill
        time.sleep(1)
        driver.close()
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(4)
        begin = begin + 1
        
    next_button = driver.find_element(By.XPATH,'//*[@id="table_pop_up_next"]')
    next.click()
    time.sleep(4)

workbook.save(filename=str(COMPANY_NAME)+".xlsx")